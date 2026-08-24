#!/usr/bin/env python3
"""
Reads JSONL (one testing_run_result-joined record per line) from stdin,
produced by export_agentic_results.js, and writes it into an .xlsx sheet
in the same column layout as the manually-built "Akamai Prompts Red Teaming.xlsx"
reference sheet ("Agentic Issues"), covering both issues and non-issues.

Usage: python3 build_xlsx.py <output.xlsx>
"""
import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter

HEADER = [
    "Issue Name", "Category", "Severity", "API Endpoint", "Testing Run",
    "Prompt 1", "Response 1",
    "Prompt 2", "Response 2",
    "Prompt 3", "Response 3",
    "Prompt 4", "Response 4",
    "Validation Message", "Result Type", "Issue Status",
]
SEVERITY_COL = "C"     # keep in sync with HEADER above
RESULT_TYPE_COL = "O"  # keep in sync with HEADER above

SEVERITY_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
RESULT_TYPE_ORDER = {"Issue": 0, "Non-Issue": 1}

# Same severity row-highlight colors as the "Akamai Prompts Red Teaming.xlsx" reference sheet.
SEVERITY_FILL = {
    "CRITICAL": "FFF8CBCB",
    "HIGH": "FFFDE2C7",
    "MEDIUM": "FFFFF2B2",
    "LOW": "FFD6EFD8",
}
HEADER_FILL = "FF4A4A4A"


def clean(value):
    if value is None:
        return ""
    s = str(value)
    return "" if s == "null" else s


def build_row(record):
    turns = record.get("turns") or []
    row = [
        record.get("name") or record.get("testSubType") or "",
        record.get("category") or "",
        record.get("severity") or "",
        f"{record.get('method', '')} {record.get('url', '')}".strip(),
        record.get("testingRun") or "",
    ]
    for i in range(4):
        if i < len(turns):
            row.append(clean(turns[i].get("finalSentPrompt")))
            row.append(clean(turns[i].get("response")))
        else:
            row.append("")
            row.append("")
    last_validation = ""
    for t in reversed(turns):
        msg = clean(t.get("validationMessage"))
        if msg:
            last_validation = msg
            break
    row.append(last_validation)
    row.append(record.get("resultType", ""))
    row.append(record.get("issueStatus") or "")
    return row


def main():
    if len(sys.argv) < 2:
        print("Usage: build_xlsx.py <output.xlsx>", file=sys.stderr)
        sys.exit(1)
    out_path = sys.argv[1]

    records = []
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        records.append(json.loads(line))

    records.sort(key=lambda r: (
        RESULT_TYPE_ORDER.get(r.get("resultType"), 9),
        0 if r.get("issueStatus") == "OPEN" else 1,
        SEVERITY_ORDER.get(r.get("severity"), 9),
        r.get("name") or r.get("testSubType") or "",
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agentic Issues"
    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(patternType="solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wrap_top = Alignment(wrap_text=True, vertical="top")
    for record in records:
        ws.append(build_row(record))
        for cell in ws[ws.max_row]:
            cell.alignment = wrap_top

    ws.freeze_panes = "A2"
    widths = [45, 22, 12, 55, 32, 45, 45, 45, 45, 45, 45, 45, 45, 55, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    last_col = get_column_letter(len(HEADER))
    last_row = max(ws.max_row, 2)
    full_range = f"A2:{last_col}{last_row}"
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Only color rows that are actual Issues -- a Non-Issue row still carries the probe's
    # inherent severity (every template has one, issue or not), so gating on severity alone
    # colored passing tests the same as failing ones. Gate on Result Type too.
    for severity, color in SEVERITY_FILL.items():
        dxf = DifferentialStyle(fill=PatternFill(patternType="solid", fgColor=color, bgColor=color))
        rule = Rule(
            type="expression",
            dxf=dxf,
            formula=[f'AND(${SEVERITY_COL}2="{severity}",${RESULT_TYPE_COL}2="Issue")'],
        )
        ws.conditional_formatting.add(full_range, rule)

    wb.save(out_path)
    counts = {}
    total_prompts = 0
    for r in records:
        rt = r.get("resultType", "Unknown")
        counts[rt] = counts.get(rt, 0) + 1
        total_prompts += sum(1 for t in (r.get("turns") or []) if clean(t.get("finalSentPrompt")))
    breakdown = ", ".join(f"{v} {k}" for k, v in sorted(counts.items(), key=lambda kv: RESULT_TYPE_ORDER.get(kv[0], 9)))
    print(f"Wrote {len(records)} rows / {total_prompts} individual prompts ({breakdown}) to {out_path}")


if __name__ == "__main__":
    main()
